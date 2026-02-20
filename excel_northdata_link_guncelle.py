#!/usr/bin/env python3
"""
Mevcut Excel dosyasındaki North Data linklerini yeni forma çevirir:
Eski: northdata.com/?q=... veya northdata.de/?q=... (ana sayfaya gidiyor)
Yeni: Google site:northdata.com araması (firma sayfasına giden sonuç)
Kullanım: python3 excel_northdata_link_guncelle.py [dosya.xlsx]  (dosya verilmezse tüm .xlsx güncellenir)
"""

import sys
from pathlib import Path
from urllib.parse import quote

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("Gerekli: pip install openpyxl")
    sys.exit(1)

LINK_FONT = Font(color="0563C1", underline="single")
NORTH_DATA_LINK_HEADER = "North_Data_Link"
FIRMA_HEADER = "Firma"
ADRESSE_HEADER = "Adresse"
HRB_HEADER = "HRB"


def _northdata_arama_metni(firma: str, adresse: str) -> str:
    """Kısa arama metni: Firma + şehir (virgülden sonrası). Tam adres çok spesifik kalıyor, sonuç çıkmıyor."""
    f = (firma or "").strip()
    adr = (adresse or "").strip()
    sehir_kismi = adr.split(",")[-1].strip() if ("," in adr) else adr
    return f"{f} {sehir_kismi}".strip() or f or sehir_kismi


def yeni_northdata_link(firma: str, adresse: str, hrb: str = "") -> str:
    """HRB varsa direkt North Data firma sayfası, yoksa Google site araması."""
    if hrb and (firma or adresse):
        sehir = adresse.split(",")[-1].strip() if (adresse and "," in adresse) else (adresse or "")
        firma_enc = quote((firma or "").strip())
        sehir_enc = quote(sehir)
        return f"https://www.northdata.com/{firma_enc},%20{sehir_enc}/{quote(hrb.strip())}"
    metin = _northdata_arama_metni(firma, adresse)
    if not metin:
        return "https://www.northdata.com/"
    return "https://www.google.com/search?q=" + quote("site:northdata.com " + metin)


def guncellenecek_link_mi(url: str) -> bool:
    """Bu link kısa formata (Firma + şehir) çevrilmeli mi? Eski ?q= veya mevcut Google araması."""
    if not url or not isinstance(url, str):
        return False
    u = url.strip()
    if "northdata.com/?q=" in u or "northdata.de/?q=" in u:
        return True
    if "google.com/search" in u and "northdata.com" in u:
        return True  # mevcut Google site araması da kısaltılsın
    return False


def excel_northdata_guncelle(dosya_yolu: str) -> bool:
    """
    Excel dosyasındaki North_Data_Link sütununu yeni forma çevirir.
    Başarılı ise True, hata/atlanan ise False.
    """
    path = Path(dosya_yolu)
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        print(f"Atlanıyor (dosya yok veya .xlsx değil): {dosya_yolu}")
        return False

    wb = load_workbook(path)
    ws = wb.active
    if ws is None:
        print(f"Atlanıyor (sayfa yok): {dosya_yolu}")
        return False

    # Başlık satırından sütun indekslerini bul
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        idx_link = headers.index(NORTH_DATA_LINK_HEADER)
        idx_firma = headers.index(FIRMA_HEADER)
        idx_adresse = headers.index(ADRESSE_HEADER)
    except ValueError as e:
        print(f"Atlanıyor ({dosya_yolu}): Gerekli sütun yok - {e}")
        return False

    col_link = idx_link + 1
    col_firma = idx_firma + 1
    col_adresse = idx_adresse + 1
    idx_hrb = headers.index(HRB_HEADER) if HRB_HEADER in headers else None
    col_hrb = (idx_hrb + 1) if idx_hrb is not None else None
    guncellenen = 0

    for row_idx in range(2, ws.max_row + 1):
        firma = (ws.cell(row=row_idx, column=col_firma).value or "").strip() or ""
        adresse = (ws.cell(row=row_idx, column=col_adresse).value or "").strip() or ""
        hrb = (ws.cell(row=row_idx, column=col_hrb).value or "").strip() if col_hrb else ""
        cell_link = ws.cell(row=row_idx, column=col_link)
        url = (cell_link.value or "").strip() if cell_link.value else ""
        # HRB doluysa direkt link; yoksa eski/uzun linkleri kısa formata çevir
        if col_hrb and hrb:
            yeni_url = yeni_northdata_link(firma, adresse, hrb)
        elif not guncellenecek_link_mi(url):
            continue
        else:
            yeni_url = yeni_northdata_link(firma, adresse)
        cell_link.value = yeni_url
        cell_link.hyperlink = yeni_url
        cell_link.font = LINK_FONT
        guncellenen += 1

    if guncellenen > 0:
        wb.save(path)
        print(f"  {path.name}: {guncellenen} North Data linki güncellendi.")
        return True
    else:
        print(f"  {path.name}: Güncellenecek eski link yok.")
        return False


def main():
    if len(sys.argv) > 1:
        dosyalar = [sys.argv[1]]
    else:
        proje = Path(__file__).resolve().parent
        dosyalar = sorted(proje.glob("turk_isletmeleri_nrw*.xlsx"))

    if not dosyalar:
        print("Hiç .xlsx dosyası bulunamadı.")
        return
    print("North Data linkleri yeni forma çevriliyor...")
    for d in dosyalar:
        excel_northdata_guncelle(str(d))
    print("Bitti.")


if __name__ == "__main__":
    main()
