#!/usr/bin/env python3
"""Veriyi CSV veya Excel (.xlsx) olarak dışa aktarır."""

import csv
from urllib.parse import quote


def csv_yaz(dosya_yolu: str, satirlar: list[dict]) -> None:
    """Liste sözlüklerini CSV dosyasına yazar."""
    if not satirlar:
        return
    alanlar = list(satirlar[0].keys())
    with open(dosya_yolu, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=alanlar)
        w.writeheader()
        w.writerows(satirlar)


# Excel'de tıklanabilir link olacak sütunlar
_LINK_SUTUNLARI = ("Google_Maps_Link", "North_Data_Link", "Website")


def google_maps_rota_url(satirlar: list[dict], max_waypoints: int = 25) -> str:
    """
    Firmaları sırayla (en yakından en uzağa) tek bir Google Maps rota linkine dönüştürür.
    Tek haritada tüm noktalar rota olarak açılır. Google'ın waypoint sınırı nedeniyle
    varsayılan en fazla 25 adres eklenir.
    """
    if not satirlar:
        return ""
    base = "https://www.google.com/maps/dir/"
    parts = []
    for row in satirlar[:max_waypoints]:
        adres = (row.get("Adresse") or "").strip()
        firma = (row.get("Firma") or "").strip()
        if firma and adres:
            search = f"{firma}, {adres}"
        else:
            search = adres or firma
        if search:
            parts.append(quote(search, safe=""))
    if not parts:
        return ""
    return base + "/".join(parts)


def excel_yaz(dosya_yolu: str, satirlar: list[dict]) -> None:
    """Liste sözlüklerini Excel (.xlsx) dosyasına yazar; link sütunları tıklanabilir hyperlink olur."""
    if not satirlar:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("Excel çıktısı için: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmalar"
    alanlar = list(satirlar[0].keys())
    ws.append(alanlar)
    link_font = Font(color="0563C1", underline="single")
    for row in satirlar:
        ws.append([row.get(k, "") for k in alanlar])
    # Link sütunlarını tıklanabilir yap (son eklenen satırlara hyperlink ver)
    for col_idx, alan in enumerate(alanlar, start=1):
        if alan not in _LINK_SUTUNLARI:
            continue
        for row_idx, row in enumerate(satirlar, start=2):
            url = (row.get(alan) or "").strip()
            if url and url.startswith("http"):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.hyperlink = url
                cell.font = link_font
                if not cell.value:
                    cell.value = url
    wb.save(dosya_yolu)


def excel_bytes(satirlar: list[dict]) -> bytes:
    """Aynı sütunlarla Excel dosyasını byte olarak döndürür (Streamlit indirme vb. için). Tıklanabilir linkler dahil."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("Excel çıktısı için: pip install openpyxl")
    bos_alanlar = ["Firma", "Adresse", "Telefon", "Website", "Region", "Sektor", "Quelle", "Aranan_Sektorler", "HRB", "Google_Maps_Link", "North_Data_Link"]
    if not satirlar:
        wb = Workbook()
        ws = wb.active
        ws.title = "Firmalar"
        ws.append(bos_alanlar)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmalar"
    alanlar = list(satirlar[0].keys())
    ws.append(alanlar)
    link_font = Font(color="0563C1", underline="single")
    for row in satirlar:
        ws.append([row.get(k, "") for k in alanlar])
    for col_idx, alan in enumerate(alanlar, start=1):
        if alan not in _LINK_SUTUNLARI:
            continue
        for row_idx, row in enumerate(satirlar, start=2):
            url = (row.get(alan) or "").strip()
            if url and url.startswith("http"):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.hyperlink = url
                cell.font = link_font
                if not cell.value:
                    cell.value = url
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def veri_yaz(dosya_yolu: str, satirlar: list[dict], alanlar: list[str] | None = None) -> None:
    """
    Veriyi dosya uzantısına göre CSV veya Excel olarak yazar.
    .xlsx / .xls -> Excel, diğerleri -> CSV.
    satirlar boş olsa bile dosya oluşturulur (başlık + 0 satır); alanlar verilmezse Firma, Adresse, ... kullanılır.
    """
    p = dosya_yolu.lower()
    if not satirlar:
        bos_alanlar = alanlar or ["Firma", "Adresse", "Telefon", "Website", "Region", "Sektor", "Quelle", "Aranan_Sektorler", "HRB", "Google_Maps_Link", "North_Data_Link"]
        if p.endswith(".xlsx") or p.endswith(".xls"):
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Firmalar"
                ws.append(bos_alanlar)
                wb.save(dosya_yolu)
            except Exception:
                pass
        else:
            with open(dosya_yolu, "w", newline="", encoding="utf-8") as f:
                import csv
                csv.DictWriter(f, fieldnames=bos_alanlar).writeheader()
        return
    if p.endswith(".xlsx") or p.endswith(".xls"):
        excel_yaz(dosya_yolu, satirlar)
    else:
        csv_yaz(dosya_yolu, satirlar)
