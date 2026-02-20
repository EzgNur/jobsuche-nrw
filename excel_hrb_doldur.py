#!/usr/bin/env python3
"""
Excel'deki boş HRB sütununu OpenRegister API ile otomatik doldurur.
Ücretsiz kota: 50 istek/ay (openregister.de ücretsiz plan).

Kullanım:
  python3 excel_hrb_doldur.py dosya.xlsx
  python3 excel_hrb_doldur.py   # turk_isletmeleri_nrw*.xlsx dosyalarından ilkini kullanır

Gereksinim: .env içinde OPENREGISTER_API_KEY (https://openregister.de/keys)
"""

import os
import sys
import time
from pathlib import Path
from urllib.parse import quote

import requests
from dotenv import load_dotenv

load_dotenv()

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("Gerekli: pip install openpyxl")
    sys.exit(1)

OPENREGISTER_API_KEY = os.getenv("OPENREGISTER_API_KEY")
API_BASE = "https://api.openregister.de/v0"
FIRMA_HEADER = "Firma"
ADRESSE_HEADER = "Adresse"
HRB_HEADER = "HRB"
NORTH_DATA_LINK_HEADER = "North_Data_Link"
LINK_FONT = Font(color="0563C1", underline="single")


def _sehir_adresden(adresse: str) -> str:
    """Adres satırından şehir (virgülden sonraki kısımdaki son kelime veya tümü)."""
    if not adresse or not isinstance(adresse, str):
        return ""
    s = adresse.strip()
    if "," in s:
        s = s.split(",")[-1].strip()  # "40215 Düsseldorf"
    return s


def openregister_ara(firma: str, api_key: str) -> list[dict]:
    """OpenRegister API ile firma ara; sonuç listesi döner (register_number, register_type, register_court, name)."""
    if not api_key or api_key == "your_api_key_here":
        return []
    url = f"{API_BASE}/search/company"
    params = {"query": (firma or "").strip(), "per_page": 10}
    headers = {"Authorization": f"Bearer {api_key}"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        return data.get("results") or []
    except requests.RequestException:
        return []
    except Exception:
        return []


def _en_uygun_sonuc(results: list[dict], firma: str, sehir: str) -> dict | None:
    """Şehir veya isim benzerliğine göre en uygun sonucu seç."""
    if not results:
        return None
    firma_lower = (firma or "").strip().lower()
    sehir_lower = (sehir or "").strip().lower()
    for r in results:
        court = (r.get("register_court") or "").lower()
        name = (r.get("name") or "").lower()
        if sehir_lower and sehir_lower in court:
            return r
        if firma_lower in name or name in firma_lower:
            return r
    return results[0]


def _hrb_metni(sonuc: dict) -> str:
    """API sonucundan North Data formatında HRB/HRA metni: 'HRB 109531'."""
    rtype = (sonuc.get("register_type") or "HRB").strip().upper()
    if not rtype:
        rtype = "HRB"
    num = (sonuc.get("register_number") or "").strip()
    if not num:
        return ""
    return f"{rtype} {num}"


def excel_hrb_doldur(dosya_yolu: str, api_key: str, gecikme_saniye: float = 0.5) -> tuple[int, int]:
    """
    Excel'de boş HRB hücrelerini OpenRegister ile doldurur.
    Döner: (doldurulan_sayisi, atlanan_sayisi)
    """
    path = Path(dosya_yolu)
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        return (0, 0)

    wb = load_workbook(path)
    ws = wb.active
    if not ws:
        return (0, 0)

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        idx_firma = headers.index(FIRMA_HEADER)
        idx_adresse = headers.index(ADRESSE_HEADER)
        idx_hrb = headers.index(HRB_HEADER)
    except ValueError:
        return (0, 0)

    col_firma = idx_firma + 1
    col_adresse = idx_adresse + 1
    col_hrb = idx_hrb + 1
    idx_link = headers.index(NORTH_DATA_LINK_HEADER) if NORTH_DATA_LINK_HEADER in headers else None
    col_link = (idx_link + 1) if idx_link is not None else None

    doldurulan = 0
    atlanan = 0

    for row_idx in range(2, ws.max_row + 1):
        hrb_cell = ws.cell(row=row_idx, column=col_hrb)
        if (hrb_cell.value or "").strip():
            atlanan += 1
            continue
        firma = (ws.cell(row=row_idx, column=col_firma).value or "").strip() or ""
        if not firma:
            atlanan += 1
            continue
        adresse = (ws.cell(row=row_idx, column=col_adresse).value or "").strip() or ""
        sehir = _sehir_adresden(adresse)

        results = openregister_ara(firma, api_key)
        time.sleep(gecikme_saniye)
        sonuc = _en_uygun_sonuc(results, firma, sehir)
        if not sonuc:
            atlanan += 1
            continue

        hrb_str = _hrb_metni(sonuc)
        if not hrb_str:
            atlanan += 1
            continue
        hrb_cell.value = hrb_str
        doldurulan += 1

        # North_Data_Link varsa direkt linke güncelle (opsiyonel)
        if col_link:
            firma_enc = quote(firma)
            sehir_enc = quote(sehir)
            yeni_url = f"https://www.northdata.com/{firma_enc},%20{sehir_enc}/{quote(hrb_str)}"
            link_cell = ws.cell(row=row_idx, column=col_link)
            link_cell.value = yeni_url
            link_cell.hyperlink = yeni_url
            link_cell.font = LINK_FONT

    if doldurulan > 0:
        wb.save(path)
    return (doldurulan, atlanan)


def main():
    if not OPENREGISTER_API_KEY or OPENREGISTER_API_KEY == "your_api_key_here":
        print("OPENREGISTER_API_KEY gerekli. .env dosyasına ekleyin.")
        print("Ücretsiz anahtar: https://openregister.de/keys (50 istek/ay)")
        sys.exit(1)

    if len(sys.argv) > 1:
        dosya = sys.argv[1]
    else:
        proje = Path(__file__).resolve().parent
        xlsxler = sorted(proje.glob("turk_isletmeleri_nrw*.xlsx"))
        if not xlsxler:
            print("Kullanım: python3 excel_hrb_doldur.py <dosya.xlsx>")
            sys.exit(1)
        dosya = xlsxler[0]

    print(f"OpenRegister ile HRB dolduruluyor: {dosya}")
    d, a = excel_hrb_doldur(dosya, OPENREGISTER_API_KEY)
    print(f"  Doldurulan: {d}, Atlanan: {a}")
    if d > 0:
        print("North Data linkleri de güncellendi (HRB dolu satırlar için direkt link).")


if __name__ == "__main__":
    main()
